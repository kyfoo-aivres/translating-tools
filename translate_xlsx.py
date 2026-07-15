from googletrans import Translator
import argparse
import os
import sys
import shutil
from tqdm import tqdm
import asyncio
from openpyxl import load_workbook

async def translate_excel(input_file, source_lang='zh-CN', target_lang='en'):
    """
    Translate Excel file from source language to target language.
    
    Args:
        input_file: Path to the input Excel file
        source_lang: Source language code (default: 'zh-CN')
        target_lang: Target language code (default: 'en')
    """
    
    # Initialize issues list
    issues = []
    
    # Check if file exists
    if not os.path.exists(input_file):
        print(f"Error: File '{input_file}' not found.")
        return
    
    # Validate file extension
    if not input_file.endswith(('.xlsx', '.xls')):
        print("Error: Input file must be an Excel file (.xlsx or .xls)")
        return
    
    # Get the base name and extension
    base_name, ext = os.path.splitext(input_file)
    
    # Create output filename by appending target language
    output_file = f"{base_name}_{target_lang}{ext}"
    
    # Copy the input file to output file
    shutil.copy2(input_file, output_file)
    print(f"Created copy: {output_file}")
    
    # Load workbook using openpyxl to preserve formatting
    workbook = load_workbook(output_file)
    sheets = workbook.sheetnames
    
    print(f"Total sheets to process: {len(sheets)}")
    print(f"Source language: {source_lang}")
    print(f"Target language: {target_lang}")
    print("-" * 50)
    
    # Initialize translator
    translator = Translator()
    
    # First pass: count total cells with content for progress bar
    total_cells = 0
    
    print("Analyzing Excel file...")
    for sheet_name in sheets:
        sheet = workbook[sheet_name]
        cell_count = 0
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip():
                    cell_count += 1
        total_cells += cell_count
        print(f"  Sheet '{sheet_name}': {cell_count} cells with content")
    
    print(f"\nTotal cells to translate: {total_cells}")
    print("Starting translation...\n")
    
    # Process each sheet
    with tqdm(total=total_cells, desc="Translating", unit="cell", 
              ncols=80, bar_format='{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]') as pbar:
        
        for sheet_name in sheets:
            sheet = workbook[sheet_name]
            pbar.set_description(f"Processing: {sheet_name}")
            
            # Iterate through each cell
            for row in sheet.iter_rows():
                for cell in row:
                    cell_value = cell.value
                    
                    # Check if cell has content (not null and not empty string)
                    if cell_value is not None and str(cell_value).strip():
                        translated = False
                        for attempt in range(3):
                            try:
                                # Translate the cell content
                                result = await translator.translate(
                                    str(cell_value), 
                                    src=source_lang, 
                                    dest=target_lang
                                )
                                if not result.text or not result.text.strip():
                                    raise ValueError("Empty or blank translation result")
                                # Replace only the text, keeping all formatting intact
                                cell.value = result.text
                                translated = True
                                break
                            except Exception as e:
                                error_str = str(e)
                                if "[Errno 11001] getaddrinfo failed" in error_str or not error_str.strip() or "Empty or blank translation result" in error_str:
                                    if attempt < 2:
                                        await asyncio.sleep(5)
                                        continue
                                # If not retryable or last attempt, fail
                                issue_msg = f"Failed to translate cell [{cell.coordinate}] after {attempt+1} attempts: {e}"
                                issues.append(issue_msg)
                                print(f"\nWarning: {issue_msg}")
                                break
                        
                        pbar.update(1)
    
    print("\nTranslating sheet names...")
    # Translate sheet names after all content is translated
    for sheet_name in sheets:
        translated_sheet = False
        for attempt in range(3):
            try:
                result = await translator.translate(
                    sheet_name, 
                    src=source_lang, 
                    dest=target_lang
                )
                if not result.text or not result.text.strip():
                    raise ValueError("Empty or blank translation result")
                translated_name = result.text
                # Rename the sheet if translation is different
                if translated_name != sheet_name:
                    workbook[sheet_name].title = translated_name
                    print(f"  Sheet '{sheet_name}' renamed to '{translated_name}'")
                translated_sheet = True
                break
            except Exception as e:
                error_str = str(e)
                if "[Errno 11001] getaddrinfo failed" in error_str or not error_str.strip() or "Empty or blank translation result" in error_str:
                    if attempt < 2:
                        await asyncio.sleep(5)
                        continue
                issue_msg = f"Failed to translate sheet name '{sheet_name}' after {attempt+1} attempts: {e}"
                issues.append(issue_msg)
                print(f"Warning: {issue_msg}")
                break
    
    # Save the workbook with all formatting preserved
    workbook.save(output_file)
    
    print("\n" + "=" * 50)
    print(f"Translation completed successfully!")
    print(f"Output file: {output_file}")
    print("=" * 50)
    
    # Save issues to file if any
    if issues:
        issues_file = f"issues_{base_name}_{target_lang}.txt"
        with open(issues_file, 'w', encoding='utf-8') as f:
            f.write("\n".join(issues))
        print(f"Issues saved to: {issues_file}")


def translate_file(input_file, source_lang, target_lang):
    """Translate a single Excel file."""
    if not input_file.lower().endswith('.xlsx'):
        print(f"Error: File '{input_file}' must be a .xlsx file.")
        return
    asyncio.run(translate_excel(input_file, source_lang, target_lang))


def main():
    """Main entry point for the script."""
    parser = argparse.ArgumentParser(
        description='Translate Excel .xlsx files while preserving formatting.'
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        '-f', '--file',
        dest='input_file',
        help='Path to a single Excel file (.xlsx)'
    )
    group.add_argument(
        '-d', '--directory',
        dest='input_dir',
        help='Path to a directory containing .xlsx files for batch translation'
    )
    parser.add_argument(
        '--source',
        default='zh-CN',
        help='Source language code (default: zh-CN)'
    )
    parser.add_argument(
        '--target',
        default='en',
        help='Target language code (default: en)'
    )

    args = parser.parse_args()

    if args.input_file:
        translate_file(args.input_file, args.source, args.target)
        return

    if not os.path.isdir(args.input_dir):
        print(f"Error: Directory '{args.input_dir}' not found.")
        sys.exit(1)

    excel_files = [
        os.path.join(args.input_dir, filename)
        for filename in sorted(os.listdir(args.input_dir))
        if filename.lower().endswith('.xlsx')
    ]

    if not excel_files:
        print(f"No .xlsx files found in '{args.input_dir}'.")
        sys.exit(1)

    for excel_file in excel_files:
        translate_file(excel_file, args.source, args.target)


if __name__ == '__main__':
    main()